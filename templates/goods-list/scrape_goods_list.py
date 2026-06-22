"""
微信小店 - 商品列表采集
调用 scanProductPreview API 获取全部商品信息
输出: JSON + Excel 到桌面

跨平台兼容: macOS / Windows / Linux
用法: python3 scrape_goods_list.py
"""

import os, json, requests, time, tempfile, sys
from collections import Counter

# ═══ 跨平台路径工具 ═══

def get_temp_file(filename):
    """跨平台临时文件路径"""
    return os.path.join(tempfile.gettempdir(), filename)

def get_desktop_path(sub_dir=None):
    """跨平台桌面路径 (macOS/Windows/Linux)"""
    home = os.path.expanduser("~")
    if sys.platform == 'darwin':  # macOS
        base = os.path.join(home, "Desktop")
    elif sys.platform == 'win32':  # Windows
        base = os.path.join(home, "Desktop")
    else:  # Linux (XDG约定)
        base = os.environ.get('XDG_DESKTOP_DIR',
                              os.path.join(home, "Desktop"))
    if not os.path.exists(base):
        base = home  # 回退到用户目录
    return os.path.join(base, sub_dir) if sub_dir else base

def get_platform_ua():
    """跨平台 User-Agent"""
    if sys.platform == 'darwin':
        return 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
    elif sys.platform == 'win32':
        return 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    else:
        return 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36'

# ═══ 配置 ═══

SKILL_DIR = os.path.dirname(os.path.abspath(__file__))
STATE_PATH = os.path.join(SKILL_DIR, '..', '..', 'assets', 'weixin_store_state.json')
OUT_DIR = get_desktop_path("微信小店商品数据")
TEMP_JSON = get_temp_file("weixin_goods_data.json")
os.makedirs(OUT_DIR, exist_ok=True)

API_URL = "https://store.weixin.qq.com/shop-faas/mmchannelstradeproductcore/cgi/goods/scanProductPreview?token=&lang=zh_CN"

STATUS_MAP = {1: '销售中', 2: '已下架', 5: '待审核', 11: '审核未通过', -1: '全部'}
HEADERS_TEMPLATE = [
    '商品ID', 'SPU编码', '商品名称', '副标题', '价格(元)', '最高价(元)', '总库存', '总销量',
    '总订单数', '总曝光量', '状态', '子状态', '上架时间', '编辑时间', '品牌', 'SKU数',
    '商品类型', '类目ID', '发货方式', '主图', '销售渠道'
]


def get_auth_headers():
    with open(STATE_PATH) as f:
        state = json.load(f)
    cookies = {c['name']: c['value'] for c in state.get('cookies', []) if c.get('name')}
    biz_magic = cookies.get('biz_magic', '')
    return {
        'Cookie': '; '.join(f'{k}={v}' for k, v in cookies.items()),
        'User-Agent': get_platform_ua(),
        'Content-Type': 'application/json',
        'Origin': 'https://store.weixin.qq.com',
        'Referer': 'https://store.weixin.qq.com/shop/goods/list',
        'biz_magic': biz_magic,
        'potter-scene': 'weixinShop',
    }


def fetch_all_products(headers):
    all_products = []
    page = 1
    while True:
        body = {
            "pageSize": 20,
            "productStatus": [-1],
            "productSource": "[1,16,32]",
            "searchSource": 1,
            "useNew": True,
            "fromProductManager": 1,
            "pageNum": page,
            "status": [-1]
        }
        resp = requests.post(API_URL, json=body, headers=headers, timeout=30)
        data = resp.json()
        
        # 检测会话过期
        if resp.status_code == 403 or data.get('respStatusCode') == 200004:
            print(f"\n❌ 登录态已过期！返回: {data.get('msg', resp.text[:100])}", flush=True)
            print(f"   请重新扫码登录并更新 assets/weixin_store_state.json\n", flush=True)
            return all_products or None
        
        products = data.get('productList', [])
        if not products:
            print(f"  ❌ 第{page}页: 无响应数据", flush=True)
            break
        all_products.extend(products)
        print(f"  第{page}页: +{len(products)} 条 (共{len(all_products)})", flush=True)
        if len(products) < 20:
            break
        page += 1
        time.sleep(0.3)
    return all_products


def extract_product_info(p):
    skus = p.get('skuList', [])
    prices = [s.get('price', 0) for s in skus if s.get('price')]
    stocks = [s.get('stockNum', 0) for s in skus]
    cat_names = [c.get('catId', '') for c in p.get('category', [])]
    return {
        '商品ID': p.get('productId', ''),
        'SPU编码': p.get('spuCode', ''),
        '商品名称': p.get('title', ''),
        '副标题': p.get('subTitle', ''),
        '价格(元)': min(prices) if prices else p.get('price', 0),
        '最高价(元)': max(prices) if prices else 0,
        '总库存': sum(stocks) if stocks else p.get('totalStockNum', 0),
        '总销量': p.get('totalSoldNum', 0),
        '总订单数': p.get('totalOrderNum', 0),
        '总曝光量': p.get('totalVisitNum', 0),
        '状态': STATUS_MAP.get(p.get('status', ''), str(p.get('status', ''))),
        '子状态': p.get('subStatus', ''),
        '上架时间': p.get('listingTime', ''),
        '编辑时间': p.get('editTime', ''),
        '品牌': p.get('brand', ''),
        'SKU数': len(skus),
        '商品类型': p.get('productType', ''),
        '类目ID': ','.join(cat_names),
        '发货方式': p.get('deliverMethod', ''),
        '主图': p.get('headImg', ''),
        '销售渠道': p.get('source', ''),
    }


def save_excel(products, fp):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品列表"

    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for ci, h in enumerate(HEADERS_TEMPLATE, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, p in enumerate(products, 2):
        info = extract_product_info(p)
        for ci, h in enumerate(HEADERS_TEMPLATE, 1):
            ws.cell(row=ri, column=ci, value=info.get(h, '')).alignment = Alignment(
                vertical="center", wrap_text=True
            )

    for ci in range(1, len(HEADERS_TEMPLATE) + 1):
        ml = max(8, len(str(HEADERS_TEMPLATE[ci - 1])))
        for ri in range(2, min(len(products) + 2, 101)):
            ml = max(ml, len(str(ws.cell(row=ri, column=ci).value or '')))
        ws.column_dimensions[get_column_letter(ci)].width = min(ml + 2, 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS_TEMPLATE))}{len(products) + 1}"
    wb.save(fp)
    return fp


def main():
    print(f"平台: {sys.platform}", flush=True)
    print("加载登录态...", flush=True)
    if not os.path.exists(STATE_PATH):
        print(f"❌ 未找到登录态文件: {STATE_PATH}")
        print("   请先在微信小店扫码登录并生成 storage_state")
        return
    headers = get_auth_headers()
    print("开始采集全部商品...", flush=True)
    products = fetch_all_products(headers)
    if products is None:
        print("❌ 登录态过期，终止采集", flush=True)
        return
    if not products:
        print("❌ 未采集到任何商品，请检查登录态是否过期", flush=True)
        return
    print(f"\n✅ 共采集 {len(products)} 件商品", flush=True)

    # JSON 备份
    with open(TEMP_JSON, 'w') as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
    print(f"📄 JSON: {TEMP_JSON}")

    # Excel
    fp = os.path.join(OUT_DIR, "商品列表_全量.xlsx")
    save_excel(products, fp)
    print(f"📄 {fp}")

    # 状态分布
    status_count = Counter(p.get('status', '') for p in products)
    print(f"\n状态分布:")
    for k, v in sorted(status_count.items()):
        print(f"  {STATUS_MAP.get(k, str(k))}: {v} 件")


if __name__ == '__main__':
    main()
