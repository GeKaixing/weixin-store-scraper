import os
import tempfile
def get_temp_file(filename):
    return os.path.join(tempfile.gettempdir(), filename)
def get_desktop_path(sub_dir=None):
    base = os.path.join(os.path.expanduser("~"), "Desktop")
    return os.path.join(base, sub_dir) if sub_dir else base

"""
两趟爬虫 (替代原 v11 one-pass):
  第一趟: 列表页 → window.open 覆盖取详情URL（不开标签，不污染Vue）→ 保存 JSON
  第二趟: 读取 JSON → 逐个开标签爬详情页 → 导出 Excel 含 26 列

优点: 两趟分离，翻页不会被 Vue 状态污染
"""
import asyncio, json, re, os
from playwright.async_api import async_playwright

SKILL_DIR = os.path.dirname(os.path.abspath(__file__))
STATE_PATH = os.path.join(SKILL_DIR, '..', '..', 'assets', 'weixin_store_state.json')
OUT_DIR = get_desktop_path("微信小店达人数据(DOM)")
TEMP_JSON = get_temp_file("weixin_pass1_data.json")

ALL_CATEGORIES = [
    '文玩文创','珠宝首饰','家纺','运动户外','母婴','家用电器','数码','鞋靴',
    '家庭清洁/纸品','箱包皮具','个人护理','食品饮料','生鲜','家居日用','家具',
    '酒类','钟表','图书','保健食品/膳食营养补充食品','服饰内衣','家装建材',
    '美妆护肤','汽摩电动','玩具乐器','教育培训','农资园艺','宠物生活',
    '成人用品','酒旅','餐饮','电脑、办公','手机通讯','厨具','其他',
]

# ═══ 修改这里 ═══
CAT_LABEL = "保健食品/膳食营养补充食品"
CAT_ID = 10000508
MAX_PAGES = 2  # 测试用 = 2, 全量 = 0 (爬全部)
FILTER_HAS_CONTACT = False  # 是否只爬有联系方式的达人
# ═════════════════

PAGINATION_INPUT_SEL = "input.weui-desktop-pagination__input"


async def toggle_category_dropdown(page):
    return await page.evaluate("""() => {
        const sr = document.querySelector('micro-app')?.shadowRoot;
        if(!sr)return false;
        for(const e of sr.querySelectorAll('*')) {
            if(e.textContent.trim()==='带货类目' && e.offsetParent!==null) {
                e.dispatchEvent(new MouseEvent('click', {bubbles:true, cancelable:true, view:window}));
                return true;
            }
        }
        return false;
    }""")


async def click_checkbox_by_label(page, label):
    return await page.evaluate(f"""((t) => {{
        const sr = document.querySelector('micro-app')?.shadowRoot;
        if(!sr)return false;
        for(const e of sr.querySelectorAll('label')){{
            if(e.textContent.trim().startsWith(t) && e.offsetParent!==null){{
                const cb = e.querySelector('input[type=checkbox]');
                if(cb) {{
                    if(cb.checked) {{
                        cb.dispatchEvent(new MouseEvent('click',{{bubbles:true,cancelable:true,view:window}}));
                        cb.dispatchEvent(new Event('change',{{bubbles:true}}));
                        cb.dispatchEvent(new Event('input',{{bubbles:true}}));
                    }}
                    cb.dispatchEvent(new MouseEvent('click',{{bubbles:true,cancelable:true,view:window}}));
                    cb.dispatchEvent(new Event('change',{{bubbles:true}}));
                    cb.dispatchEvent(new Event('input',{{bubbles:true}}));
                }}
                return true;
            }}
        }}
        return false;
    }})('{label}')""")


async def toggle_contact_filter(page, enabled=True):
    """勾选/取消 '有联系方式' 筛选 checkbox"""
    flag = "true" if enabled else "false"
    return await page.evaluate(f"""((en) => {{
        const sr = document.querySelector('micro-app')?.shadowRoot;
        if(!sr)return false;
        const cb = sr.querySelector('input[value=hasContact]');
        if(!cb) return false;
        const wantCheck = en;
        if(cb.checked === wantCheck) return true;
        cb.dispatchEvent(new MouseEvent('click',{{bubbles:true,cancelable:true,view:window}}));
        cb.dispatchEvent(new Event('change',{{bubbles:true}}));
        cb.dispatchEvent(new Event('input',{{bubbles:true}}));
        return true;
    }})({flag})""")


async def get_max_page(page):
    return await page.evaluate("""() => {
        const sr = document.querySelector('micro-app')?.shadowRoot;
        if(!sr)return 0;
        const pag = sr.querySelector('[class*="pagination"]');
        if(!pag)return 0;
        const text = pag.textContent;
        let m = text.match(/\.\.\.(\d+)/);
        if(m) return parseInt(m[1]);
        m = text.match(/共(\d+)页/);
        if(m) return parseInt(m[1]);
        const nums = text.match(/\d+/g);
        if(nums) return Math.max(...nums.map(Number));
        return 0;
    }""")


async def goto_page_via_input(page, num):
    """翻页: native value setter + input/change 事件 → dispatchEvent 点跳转"""
    ok = await page.evaluate(f"""((n) => {{
        const sr = document.querySelector('micro-app')?.shadowRoot;
        if(!sr)return false;
        const inp = sr.querySelector('{PAGINATION_INPUT_SEL}');
        if(!inp) return false;
        const setter = Object.getOwnPropertyDescriptor(
            window.HTMLInputElement.prototype, 'value'
        ).set;
        setter.call(inp, String(n));
        inp.dispatchEvent(new Event('input', {{bubbles:true}}));
        inp.dispatchEvent(new Event('change', {{bubbles:true}}));
        for(const e of sr.querySelectorAll('a')) {{
            if(e.textContent.trim()==='跳转') {{
                e.dispatchEvent(new MouseEvent('click', {{bubbles:true,cancelable:true,view:window}}));
                return true;
            }}
        }}
        return false;
    }})({num})""")
    await asyncio.sleep(3)
    return ok


async def get_list_data_with_urls(page):
    """获取列表数据 + window.open 覆盖捕获详情URL"""
    data = await page.evaluate("""() => {
        const sr = document.querySelector('micro-app')?.shadowRoot;
        if(!sr)return {rows:[]};
        const rows = [];
        for(const tr of sr.querySelectorAll('tr')){
            const t=tr.textContent.trim();
            if(!t||t.includes('达人昵称'))continue;
            const tds=tr.querySelectorAll('td');
            if(tds.length<3)continue;
            const row = Array.from(tds).map(td => td.textContent.trim());
            const img = tds[0].querySelector('img');
            rows.push({row, avatar: img ? img.src : ''});
        }
        const links = sr.querySelectorAll('td:last-child a');
        const captured = {};
        const origOpen = window.open;
        window.open = function(url) {
            captured._openUrl = url.startsWith('http') ? url
                : 'https://store.weixin.qq.com' + url;
            return {closed: false, close: function(){ this.closed=true; }};
        };
        for(let i=0;i<links.length;i++){
            try {
                links[i].dispatchEvent(new MouseEvent('click', {bubbles:true, cancelable:true, view:window}));
                if(rows[i]) { rows[i].url = captured._openUrl || ''; captured._openUrl = ''; }
            } catch(e) { if(rows[i]) rows[i].url = ''; }
        }
        window.open = origOpen;
        return {rows};
    }""")
    return data['rows']


def parse_list_row(item):
    td = item.get('row', [])
    info = td[0] if len(td) > 0 else ''
    info = re.sub(r'暂无评分[^ ]*', '', info)
    score_m = re.search(r'评分:\s*([\d.]+)', info)
    score = score_m.group(1) if score_m else ''
    badges = [b for b in ['回复率高','有认证','可开发票','近期选品活跃'] if b in info]
    cleaned = re.sub(r'评分:\s*[\d.]+', '', info)
    for b in badges: cleaned = cleaned.replace(b, '')
    cleaned = cleaned.strip()
    found_cats = []; remaining = cleaned
    for cn in sorted(ALL_CATEGORIES, key=len, reverse=True):
        if cn in remaining: found_cats.append(cn); remaining = remaining.replace(cn, '')
    nickname = remaining.strip().rstrip(',').strip()
    return {
        "头像": item.get('avatar', ''),
        "昵称": nickname,
        "带货类目": ', '.join(found_cats),
        "评分": score,
        "粉丝数(列表)": td[1] if len(td) > 1 else '',
        "带货销售总额": td[2] if len(td) > 2 else '',
        "短视频销售额": td[4] if len(td) > 4 else '',
        "回复率高": "是" if '回复率高' in badges else "否",
        "有认证": "是" if '有认证' in badges else "否",
        "可开发票": "是" if '可开发票' in badges else "否",
        "达人详情链接": item.get('url', ''),
    }


async def pass1(ctx, page):
    """第一趟: 列表数据 + 详情URL捕获"""
    print(f"\n=== 第一趟: {CAT_LABEL} (列表) ===", flush=True)
    await asyncio.sleep(4)

    for attempt in range(3):
        print(f"  筛选尝试 {attempt+1}/3...", flush=True)
        await toggle_category_dropdown(page)
        await asyncio.sleep(3)
        await click_checkbox_by_label(page, CAT_LABEL)
        await asyncio.sleep(1.5)
        await toggle_category_dropdown(page)
        if FILTER_HAS_CONTACT:
            await toggle_contact_filter(page, True)
            await asyncio.sleep(1.5)
        max_page = 0
        for _ in range(10):
            await asyncio.sleep(3)
            max_page = await get_max_page(page)
            if max_page > 0: break
        print(f"  总页数: {max_page}", flush=True)
        if max_page > 0: break
        if attempt < 2:
            print("  筛选未生效，刷新重试...", flush=True)
            await page.goto("https://store.weixin.qq.com/shop/findersquare/find",
                             wait_until="domcontentloaded", timeout=30_000)
            await asyncio.sleep(8)

    if max_page == 0:
        print("❌ 筛选失败", flush=True); return []

    all_talents = []
    prev_uniq = 0
    stale = 0
    limit = min(max_page, MAX_PAGES) if MAX_PAGES > 0 else max_page

    for pn in range(1, limit + 1):
        if pn > 1:
            await goto_page_via_input(page, pn)
        list_rows = await get_list_data_with_urls(page)
        for item in list_rows:
            all_talents.append(parse_list_row(item))
        cur = len(set(t['昵称'] for t in all_talents))
        stale = 0 if cur > prev_uniq else stale + 1
        prev_uniq = cur
        if stale >= 10:
            print(f"  P{pn}: ⛔ 连续{stale}页无新数据", flush=True); break
        has_url = sum(1 for t in all_talents[-len(list_rows):] if t.get('达人详情链接'))
        print(f"  P{pn}/{limit}: +{len(list_rows)}条 | {cur}不重复 | {has_url}链接", flush=True)
        with open(TEMP_JSON, 'w') as f:
            json.dump(all_talents, f, ensure_ascii=False)

    print(f"第一趟完成: {len(all_talents)}条, {len(set(t['昵称'] for t in all_talents))}不重复", flush=True)
    return all_talents


DETAIL_PARSE_JS = """
() => {
    function getText(el) {
        if (!el || el.tagName === 'STYLE' || el.tagName === 'SCRIPT') return '';
        let text = '';
        for (const node of el.childNodes) {
            if (node.nodeType === 3) { const t = node.textContent.trim(); if (t) text += t + '\\n'; }
            else if (node.nodeType === 1) text += getText(node);
        }
        return text;
    }
    const sr = document.querySelector('micro-app')?.shadowRoot;
    const root = sr || document;
    const body = root.querySelector('body') || root;
    let raw = getText(body);
    // fallback: 如果getText返回行太少, 用document.body.innerText
    const lines1 = raw.split('\\n').map(l => l.trim()).filter(l => l);
    if (lines1.length < 10) {
        raw = document.body.innerText || '';
    }
    const lines = raw.split('\\n').map(l => l.trim()).filter(l => l);
    function valAfter(label) {
        for (let i = 0; i < lines.length; i++)
            if (lines[i] === label && i + 1 < lines.length) return lines[i + 1];
        return '';
    }
    const 总销量 = valAfter('总销量');
    const 跟买人数 = valAfter('跟买人数');
    const 回头客 = valAfter('回头客');
    let 品类占比 = '';
    { const si = lines.indexOf('品类占比');
      if (si !== -1) {
        const pairs = [];
        for (let i = si + 1; i < lines.length; i++) {
          const l = lines[i];
          if (['带货概览','粉丝特征','直播带货','短视频带货','带货渠道'].includes(l)) break;
          if (i + 1 < lines.length && lines[i+1].endsWith('%')) {
            pairs.push(l + ' ' + lines[i+1]); i++;
          }
        }
        品类占比 = pairs.join(', ');
      }
    }
    // 带货概览: 标签和价值分行 (带货销售额 / ￥50万-100万)
    let 带货销售额='', 客单价='', 粉丝数带货概览='', 直播占比='';
    { const si = lines.indexOf('带货概览');
      if (si !== -1) {
        for (let i = si + 1; i < lines.length && i < si + 20; i++) {
          const l = lines[i];
          if (['粉丝特征','品类占比','带货渠道','直播带货','短视频带货'].includes(l)) break;
          if (l === '带货销售额' && i+1 < lines.length) 带货销售额 = lines[i+1];
          else if (l === '客单价' && i+1 < lines.length) 客单价 = lines[i+1];
          else if (l === '粉丝数' && i+1 < lines.length) 粉丝数带货概览 = lines[i+1];
          else if (l === '直播' && i+1 < lines.length) 直播占比 = lines[i+1];
        }
      }
    }
    // 粉丝特征: 标签/价值/占/百分比 四行一组
    let 粉丝性别='', 粉丝年龄='', 粉丝地域='', 粉丝人群类别='', 粉丝购物偏好='', 购买力区间='';
    { const si = lines.indexOf('粉丝特征');
      if (si !== -1) {
        for (let i = si + 1; i < lines.length && i < si + 30; i++) {
          const l = lines[i];
          if (['直播带货','带货渠道','品类占比','带货概览'].includes(l)) break;
          if (l === '性别' && i+3 < lines.length) 粉丝性别 = lines[i+1] + ' ' + lines[i+3];
          else if (l === '年龄' && i+3 < lines.length) 粉丝年龄 = lines[i+1] + ' ' + lines[i+3];
          else if (l === '地域' && i+3 < lines.length) 粉丝地域 = lines[i+1] + ' ' + lines[i+3];
          else if (l === '人群类别' && i+3 < lines.length) 粉丝人群类别 = lines[i+1] + ' ' + lines[i+3];
          else if (l === '购物偏好' && i+3 < lines.length) 粉丝购物偏好 = lines[i+1] + ' ' + lines[i+3];
          else if (l === '购买力区间' && i+3 < lines.length) 购买力区间 = lines[i+1] + ' ' + lines[i+3];
        }
      }
    }
    let 带货渠道 = '';
    { const si = lines.indexOf('带货渠道');
      if (si !== -1) {
        const kept = [];
        const skip = ['微信扫码','联系','打开微信','扫码入群','邀请带货','恭喜',
          '互相感兴趣','未留下联系方式','昨日','近7日','近30日','直播明细','直播信息',
          '直播观看人数','带货商品','带货片段','下一页','跳转','评分','达人详情','找达人',
          '带货概览','粉丝特征','品类占比','直播带货','短视频带货'];
        for (let i = si + 1; i < lines.length; i++) {
          const l = lines[i];
          if (skip.some(p => l.includes(p))) break;
          if (l.length > 1 && !/^\\\\d+$/.test(l) && !l.includes('%') && !l.startsWith('￥')) kept.push(l);
        }
        带货渠道 = kept.join('; ');
      }
    }
    // 直播带货: 场均成交额, 场均观看人数, 总带货场次 (在直播带货section下)
    let 场均成交额='', 场均观看人数='', 总带货场次='', 直播销售额='';
    { const si = lines.indexOf('直播带货');
      if (si !== -1) {
        for (let i = si + 1; i < lines.length && i < si + 30; i++) {
          const l = lines[i];
          if (['粉丝特征','带货渠道','品类占比','带货概览','短视频明细','直播明细'].includes(l)) break;
          if (l === '直播销售额' && i+1 < lines.length) 直播销售额 = lines[i+1];
          else if (l === '场均成交额' && i+1 < lines.length) 场均成交额 = lines[i+1];
          else if (l === '场均观看人数' && i+1 < lines.length) 场均观看人数 = lines[i+1];
          else if (l === '总带货场次' && i+1 < lines.length) 总带货场次 = lines[i+1];
        }
      }
    }
    // 直播明细: 解析带货片段行组 [标题/日期/时长/观看人数/等/N/件]
    let 直播明细 = '';
    { const si = lines.indexOf('直播明细');
      if (si !== -1) {
        const rows = [];
        let i = si + 1;
        // 跳过表头 直播信息/直播观看人数/带货商品/带货片段
        while (i < lines.length && ['直播信息','直播观看人数','带货商品','带货片段'].includes(lines[i])) i++;
        while (i < lines.length - 6) {
          const l = lines[i];
          if (['带货片段','下一页','跳转','直播带货商品','暂无权限'].includes(l) || l.startsWith('直播带货') || l.startsWith('短视频')) break;
          // 5字段行组: 标题 / 日期 / 时长 / 观看人数 / "等" / N / "件"
          const date = lines[i+1] || '';
          const dur = lines[i+2] || '';
          const views = lines[i+3] || '';
          const n = lines[i+5] || '';
          if (/^\\d/.test(date.charAt(0)) && dur.includes('分') && views) {
            rows.push(l + '|' + date + '|' + dur + '|' + views + '|' + n + '件');
            i += 7;
          } else { i++; }
        }
        直播明细 = rows.join('; ');
      }
    }
    return {总销量,跟买人数,回头客,品类占比,带货销售额,客单价,粉丝数带货概览,直播占比,
            粉丝性别,粉丝年龄,粉丝地域,粉丝人群类别,粉丝购物偏好,购买力区间,带货渠道,
            直播销售额,场均成交额,场均观看人数,总带货场次,直播明细};
}
"""


async def pass2():
    """第二趟: 逐个开标签爬详情数据"""
    print(f"\n=== 第二趟: 详情页爬取 ===", flush=True)
    if not os.path.exists(TEMP_JSON):
        print("❌ 无第一趟数据", flush=True); return []
    with open(TEMP_JSON) as f:
        all_talents = json.load(f)
    total = len(all_talents)
    urls = [(i, t['达人详情链接']) for i, t in enumerate(all_talents) if t.get('达人详情链接')]
    print(f"  共{total}条, {len(urls)}个详情链接", flush=True)
    if not urls:
        print("无详情链接", flush=True); return all_talents

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(storage_state=STATE_PATH, viewport={"width":1920,"height":1080})
        for idx, (orig_idx, url) in enumerate(urls):
            full_url = url if url.startswith('http') else 'https://store.weixin.qq.com' + url
            try:
                dp = await ctx.new_page()
                await dp.goto(full_url, wait_until="domcontentloaded", timeout=20000)
                await asyncio.sleep(3)
                data = await dp.evaluate(DETAIL_PARSE_JS)
                # 点击短视频带货tab(用坐标点击, dispatchEvent对channel-tab不生效)
                sv_tab = await dp.evaluate("""() => {
                    const sr = document.querySelector('micro-app')?.shadowRoot;
                    if(!sr) return null;
                    for(const e of sr.querySelectorAll('div')) {
                        if(e.offsetParent === null) continue;
                        if(e.textContent.trim() === '短视频带货') {
                            const r = e.getBoundingClientRect();
                            return {x: r.x + r.width/2, y: r.y + r.height/2};
                        }
                    }
                    return null;
                }""")
                if sv_tab:
                    await dp.mouse.click(sv_tab['x'], sv_tab['y'])
                await asyncio.sleep(3)
                # 捕获 roomId: 点击"联系"按钮 → 捕获 popup URL 中的 roomId
                room_id = ''
                popup_future = asyncio.get_event_loop().create_future()
                def on_page(new_page):
                    if not popup_future.done():
                        popup_future.set_result(new_page)
                ctx.on('page', on_page)

                contact_coord = await dp.evaluate("""() => {
                    const sr = document.querySelector('micro-app')?.shadowRoot;
                    if(!sr) return null;
                    const btns = sr.querySelectorAll('button');
                    for(const b of btns) {
                        if(b.textContent.trim() === '联系' && b.offsetParent !== null) {
                            const r = b.getBoundingClientRect();
                            return {x: r.x + r.width/2, y: r.y + r.height/2};
                        }
                    }
                    return null;
                }""")
                if contact_coord:
                    await dp.mouse.click(contact_coord['x'], contact_coord['y'])
                    try:
                        popup_page = await asyncio.wait_for(popup_future, timeout=10)
                        pu = popup_page.url
                        if 'roomId=' in pu:
                            room_id = pu.split('roomId=')[-1].split('&')[0]
                            print(f"    roomId={room_id}", end='', flush=True)
                        await popup_page.close()
                    except asyncio.TimeoutError:
                        pass

                sv_data = await dp.evaluate("""() => {
                    function getText(el) {
                        if (!el || el.tagName === 'STYLE' || el.tagName === 'SCRIPT') return '';
                        let text = '';
                        for (const node of el.childNodes) {
                            if (node.nodeType === 3) { const t = node.textContent.trim(); if (t) text += t + '\\n'; }
                            else if (node.nodeType === 1) text += getText(node);
                        }
                        return text;
                    }
                    const sr = document.querySelector('micro-app')?.shadowRoot;
                    const root = sr || document;
                    const body = root.querySelector('body') || root;
                    let raw = getText(body);
                    const chk = raw.split('\\n').map(l => l.trim()).filter(l => l);
                    if (chk.length < 10) raw = document.body.innerText || '';
                    const lines = raw.split('\\n').map(l => l.trim()).filter(l => l);
                    let 视频销售额='', 条均成交额='', 条均点赞数='', 总带货条数='', 短视频明细='';
                    const si = lines.indexOf('短视频带货');
                    if (si !== -1) {
                        for (let i = si; i < lines.length && i < si + 60; i++) {
                            const l = lines[i];
                            if (l === '视频销售额' && i+1 < lines.length) 视频销售额 = lines[i+1];
                            else if (l === '条均成交额' && i+1 < lines.length) 条均成交额 = lines[i+1];
                            else if (l === '条均点赞数' && i+1 < lines.length) 条均点赞数 = lines[i+1];
                            else if (l === '总带货条数' && i+1 < lines.length) 总带货条数 = lines[i+1];
                        }
                    }
                    // 短视频明细: 9行一组: 标题/日期/扫码看视频/点赞/分享/喜欢/商品名/价格/扫码看详情
                    { const di = lines.indexOf('短视频明细');
                      if (di !== -1) {
                        const rows = [];
                        let i = di + 1;
                        // 跳过表头: 短视频信息, 点赞数, 分享数, 喜欢数, 带货商品
                        while (i < lines.length && ['短视频信息','点赞数','分享数','喜欢数','带货商品'].includes(lines[i])) i++;
                        while (i < lines.length - 8) {
                          const l = lines[i];
                          // 结束条件
                          if (l === '暂无数据' || l.startsWith('直播') || l === '带货片段' || l === '下一页' || l === '跳转') break;
                          const date = lines[i+1] || '';
                          const likes = lines[i+3] || '';
                          const shares = lines[i+4] || '';
                          const favs = lines[i+5] || '';
                          // 检测: 第2行应该是日期(YYYY/MM/DD), 第4行应该是数字(点赞)
                          if (/^\d{4}\//.test(date) && /^\d+$/.test(likes)) {
                            rows.push(l + '|' + date + '|赞' + likes + '|分享' + shares + '|喜欢' + favs);
                            i += 9;  // 每条占9行
                          } else { i++; }
                        }
                        短视频明细 = rows.join('; ');
                      }
                    }
                    return {视频销售额,条均成交额,条均点赞数,总带货条数,短视频明细};
                }""")
                for k, v in sv_data.items():
                    data[k] = v
                for k, v in data.items():
                    all_talents[orig_idx][k] = v
                all_talents[orig_idx]['roomId'] = room_id
                all_talents[orig_idx]['imLink'] = f'https://store.weixin.qq.com/shop/kf/collab/im?mode=business&roomId={room_id}' if room_id else ''
                await dp.close()
            except Exception as e:
                print(f"  [{idx+1}/{len(urls)}] ✗ {str(e)[:50]}", end='', flush=True)
            if (idx + 1) % 10 == 0:
                with open(TEMP_JSON, 'w') as f:
                    json.dump(all_talents, f, ensure_ascii=False)
            print(f"  [{idx+1}/{len(urls)}] {'✓' if all_talents[orig_idx].get('总销量') else '△'} "
                  f"{all_talents[orig_idx]['昵称'][:18]}", flush=True)
        await browser.close()
    done = sum(1 for t in all_talents if t.get('总销量'))
    print(f"第二趟完成: {done}/{total} 有详情数据", flush=True)
    return all_talents


def save_excel(all_talents, cat_label):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    os.makedirs(OUT_DIR, exist_ok=True)
    safe_name = cat_label.replace('/', '_')
    fp = os.path.join(OUT_DIR, f"达人_带货类目_{safe_name}_含详情.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cat_label[:31].replace('/', '、').replace('\\', '、')

    # 固定列序 (41列)
    hdrs = ['头像','昵称','带货类目1','带货类目2','带货类目3','评分','粉丝数(列表)',
            '带货销售总额','短视频销售额','回复率高','有认证','可开发票','达人详情链接',
            '总销量','跟买人数','回头客','品类占比','带货销售额','客单价','粉丝数带货概览',
            '直播占比','粉丝性别','粉丝年龄','粉丝地域','粉丝人群类别','粉丝购物偏好',
            '购买力区间','带货渠道','直播销售额','场均成交额','场均观看人数','总带货场次',
            '直播明细','视频销售额','条均成交额','条均点赞数','总带货条数','短视频明细',
            '微信号','手机号','im链接']

    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11); c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, t in enumerate(all_talents, 2):
        # 把带货类目拆成3列
        cats = [c.strip() for c in t.get('带货类目', '').split(',')]
        cat1 = cats[0] if len(cats) > 0 else ''
        cat2 = cats[1] if len(cats) > 1 else ''
        cat3 = cats[2] if len(cats) > 2 else ''
        # 取值映射
        vals = {
            '头像': t.get('头像', ''),
            '昵称': t.get('昵称', ''),
            '带货类目1': cat1, '带货类目2': cat2, '带货类目3': cat3,
            '评分': t.get('评分', ''),
            '粉丝数(列表)': t.get('粉丝数(列表)', ''),
            '带货销售总额': t.get('带货销售总额', ''),
            '短视频销售额': t.get('短视频销售额', ''),
            '回复率高': t.get('回复率高', ''),
            '有认证': t.get('有认证', ''),
            '可开发票': t.get('可开发票', ''),
            '达人详情链接': t.get('达人详情链接', ''),
            '总销量': t.get('总销量', ''),
            '跟买人数': t.get('跟买人数', ''),
            '回头客': t.get('回头客', ''),
            '品类占比': t.get('品类占比', ''),
            '带货销售额': t.get('带货销售额', ''),
            '客单价': t.get('客单价', ''),
            '粉丝数带货概览': t.get('粉丝数带货概览', ''),
            '直播占比': t.get('直播占比', ''),
            '粉丝性别': t.get('粉丝性别', ''),
            '粉丝年龄': t.get('粉丝年龄', ''),
            '粉丝地域': t.get('粉丝地域', ''),
            '粉丝人群类别': t.get('粉丝人群类别', ''),
            '粉丝购物偏好': t.get('粉丝购物偏好', ''),
            '购买力区间': t.get('购买力区间', ''),
            '带货渠道': t.get('带货渠道', ''),
            '直播销售额': t.get('直播销售额', ''),
            '场均成交额': t.get('场均成交额', ''),
            '场均观看人数': t.get('场均观看人数', ''),
            '总带货场次': t.get('总带货场次', ''),
            '直播明细': t.get('直播明细', ''),
            '视频销售额': t.get('视频销售额', ''),
            '条均成交额': t.get('条均成交额', ''),
            '条均点赞数': t.get('条均点赞数', ''),
            '总带货条数': t.get('总带货条数', ''),
            '短视频明细': t.get('短视频明细', ''),
            '微信号': t.get('微信号', ''),
            '手机号': t.get('手机号', ''),
            'im链接': t.get('imLink', ''),
        }
        for ci, h in enumerate(hdrs, 1):
            ws.cell(row=ri, column=ci, value=vals.get(h, '')).alignment = Alignment(
                vertical="center", wrap_text=True
            )
    for ci, h in enumerate(hdrs, 1):
        ml = max(8, len(str(h)))
        for ri in range(2, min(len(all_talents)+2, 101)):
            ml = max(ml, len(str(ws.cell(row=ri, column=ci).value or '')))
        ws.column_dimensions[get_column_letter(ci)].width = min(ml+2, 60)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}{len(all_talents)+1}"
    wb.save(fp)
    print(f"📄 {fp}", flush=True)
    return fp


async def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(storage_state=STATE_PATH, viewport={"width":1920,"height":1080})
        page = await ctx.new_page()
        await page.goto("https://store.weixin.qq.com/shop/findersquare/find",
                         wait_until="domcontentloaded", timeout=30_000)
        await asyncio.sleep(5)
        body = await page.text_content("body") or ""
        if "登录" in body and "扫码" in body:
            print("❌ 会话过期！", flush=True)
            await browser.close(); return
        talents = await pass1(ctx, page)
        await browser.close()
        if talents:
            with open(TEMP_JSON, 'w') as f:
                json.dump(talents, f, ensure_ascii=False)
        else:
            return
    talents = await pass2()
    if talents:
        fp = save_excel(talents, CAT_LABEL)
        print(f"\n✅ {fp}", flush=True)

if __name__ == '__main__':
    asyncio.run(main())
