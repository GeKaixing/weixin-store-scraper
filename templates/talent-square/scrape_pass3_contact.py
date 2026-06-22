"""
Pass3 — 联系方式提取
从已保存的 pass1+pass2 JSON 数据中读取 roomId，
逐个打开 collab/im 页面，点击"查看联系方式"提取 wechat + phone。

输入: 读取 get_temp_file('weixin_pass1_data.json') 中的达人数据
      （要求包含 roomId 字段，由 pass2 写入）
输出: JSON + Excel (新增两列: 微信号, 手机号)

用法:
  1. 先运行 pass1+pass2 完成详情数据采集
  2. 运行本脚本 python3 scrape_pass3_contact.py
"""

import os, sys, json, asyncio, re
import tempfile

def get_temp_file(filename):
    \"\"\"跨平台临时文件路径\"\"\"
    return os.path.join(tempfile.gettempdir(), filename)

def get_desktop_path(sub_dir=None):
    \"\"\"跨平台桌面路径 (macOS/Windows/Linux)\"\"\"
    home = os.path.expanduser("~")
    if sys.platform == 'darwin':
        base = os.path.join(home, "Desktop")
    elif sys.platform == 'win32':
        base = os.path.join(home, "Desktop")
    else:
        base = os.environ.get('XDG_DESKTOP_DIR', os.path.join(home, "Desktop"))
    if not os.path.exists(base):
        base = home
    return os.path.join(base, sub_dir) if sub_dir else base

from playwright.async_api import async_playwright

SKILL_DIR = os.path.dirname(os.path.abspath(__file__))
STATE_PATH = os.path.join(SKILL_DIR, '..', '..', 'assets', 'weixin_store_state.json')
OUT_DIR = get_desktop_path("微信小店达人数据(DOM)")
TEMP_JSON = get_temp_file("weixin_pass1_data.json")
CONTACT_JSON = get_temp_file("weixin_contact_progress.json")


async def extract_contact(page, room_id, nickname):
    """从 collab/im 页面提取联系方式"""
    url = f'https://store.weixin.qq.com/shop/kf/collab/im?mode=business&roomId={room_id}'
    try:
        await page.goto(url, wait_until='domcontentloaded', timeout=20000)
        await asyncio.sleep(5)

        # 第一步: 点击左侧会话列表匹配的会话，加载右侧面板
        await page.evaluate(f"""((rid) => {{
            const items = document.querySelectorAll('[data-room-id]');
            for(const item of items) {{
                if(item.getAttribute('data-room-id') === rid) {{
                    item.click();
                    return;
                }}
            }}
            // fallback: 找第一个可见会话
            const first = document.querySelector('li[class*=\"session\"]');
            if(first) first.click();
        }})('{room_id}')""")
        await asyncio.sleep(5)

        # 第二步: 点"查看联系方式"
        clicked = await page.evaluate("""() => {
            const link = document.querySelector('a.contact-link, .contact-link, [class*="contact-link"]');
            if (link) { link.click(); return true; }
            return false;
        }""")
        if clicked:
            await asyncio.sleep(3)

        # 第三步: 提取联系方式 - 仅从 contact-popover 提取
        wechat = ''
        phone = ''
        skip = False
        contact = await page.evaluate("""() => {
            const result = {wechat: '', phone: '', skip: false, reason: ''};
            const title = document.querySelector('.contact-popover__title');
            const body = document.body.innerText || '';
            
            if (body.includes('已达上限')) {
                result.skip = true;
                result.reason = '已达上限';
                return result;
            }
            if (body.includes('暂无联系方式')) {
                result.reason = '暂无';
                return result;
            }
            
            if (title && title.textContent.trim().includes('带货者联系方式')) {
                const items = document.querySelectorAll('.contact-popover__item');
                items.forEach(item => {
                    const iconHtml = item.innerHTML;
                    const valueEl = item.querySelector('.contact-popover__value');
                    if (!valueEl) return;
                    const val = valueEl.textContent.trim();
                    if (iconHtml.includes('wechat') || iconHtml.includes('weixin')) {
                        result.wechat = val;
                    } else {
                        const pm = val.match(/1[3-9]\\d{9}/);
                        if (pm) result.phone = pm[0];
                    }
                });
            }
            return result;
        }""")
        if contact.get('skip'):
            skip = True
            print(f"  {nickname}: ⏭ 已达上限，跳过", flush=True)
        elif contact.get('reason') == '暂无':
            print(f"  {nickname}: 暂无联系方式", flush=True)
        else:
            wechat = contact.get('wechat', '')
            phone = contact.get('phone', '')
            if wechat or phone:
                print(f"  {nickname}: 📱{phone} 💬{wechat}", flush=True)
            else:
                print(f"  {nickname}: 无联系方式数据", flush=True)

        return wechat, phone, skip

    except Exception as e:
        print(f"  {nickname}: ✗ {str(e)[:60]}", flush=True)
        return '', '', False


async def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    if not os.path.exists(TEMP_JSON):
        print(f"❌ 未找到 pass1+pass2 数据: {TEMP_JSON}", flush=True)
        return

    with open(TEMP_JSON) as f:
        talents = json.load(f)
    print(f"加载 {len(talents)} 条达人数据", flush=True)

    # 加载已有联系方式进度
    done = {}
    if os.path.exists(CONTACT_JSON):
        with open(CONTACT_JSON) as f:
            done = json.load(f)
        # 合并已有数据
        for d in talents:
            nick = d.get('昵称', '')
            if nick in done:
                d['微信号'] = done[nick].get('微信号', d.get('微信号', ''))
                d['手机号'] = done[nick].get('手机号', d.get('手机号', ''))

    # 筛选需要处理 roomId 的条目
    to_process = [t for t in talents
                  if t.get('roomId') and not t.get('微信号') and not t.get('手机号')]
    if not to_process:
        # 还有没有 roomId 的？试试从 imLink 提取
        for t in talents:
            if not t.get('roomId') and t.get('imLink'):
                m = re.search(r'roomId=(\d+)', t.get('imLink', ''))
                if m:
                    t['roomId'] = m.group(1)
        to_process = [t for t in talents
                      if t.get('roomId') and not t.get('微信号') and not t.get('手机号')]

    if not to_process:
        print("✅ 没有需要提取联系方式的条目", flush=True)
        # 直接输出 Excel
        save_excel(talents)
        return

    print(f"需要提取联系方式: {len(to_process)} 条", flush=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(storage_state=STATE_PATH, viewport={"width":1920, "height":1080})

        for idx, t in enumerate(to_process):
            page = await ctx.new_page()
            try:
                wechat, phone, skip = await extract_contact(
                    page, t['roomId'], t.get('昵称', '')
                )
            finally:
                await page.close()
            if skip:
                continue
            t['微信号'] = wechat
            t['手机号'] = phone

            # 增量保存进度: 每条立刻写入 TEMP_JSON + CONTACT_JSON
            done[t.get('昵称', '')] = {'微信号': wechat, '手机号': phone}
            with open(CONTACT_JSON, 'w') as f:
                json.dump(done, f, ensure_ascii=False)
            with open(TEMP_JSON, 'w') as f:
                json.dump(talents, f, ensure_ascii=False)

        await browser.close()

    # 保存最终 JSON
    with open(TEMP_JSON, 'w') as f:
        json.dump(talents, f, ensure_ascii=False)

    # 输出 Excel
    save_excel(talents)
    print(f"✅ 完成! 共提取 {len(to_process)} 条联系方式", flush=True)


def save_excel(talents):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    os.makedirs(OUT_DIR, exist_ok=True)
    fp = os.path.join(OUT_DIR, "达人_含联系方式.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "达人数据"

    # 固定列序 (41列)
    hdrs = ['头像','昵称','带货类目1','带货类目2','带货类目3','评分','粉丝数(列表)',
            '带货销售总额','短视频销售额','回复率高','有认证','可开发票','达人详情链接',
            '总销量','跟买人数','回头客','品类占比','带货销售额','客单价','粉丝数带货概览',
            '直播占比','粉丝性别','粉丝年龄','粉丝地域','粉丝人群类别','粉丝购物偏好',
            '购买力区间','带货渠道','直播销售额','场均成交额','场均观看人数','总带货场次',
            '直播明细','视频销售额','条均成交额','条均点赞数','总带货条数','短视频明细',
            '微信号','手机号','im链接']

    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, t in enumerate(talents, 2):
        # 把带货类目拆成3列
        cats = [c.strip() for c in t.get('带货类目', '').split(',')]
        cat1 = cats[0] if len(cats) > 0 else ''
        cat2 = cats[1] if len(cats) > 1 else ''
        cat3 = cats[2] if len(cats) > 2 else ''
        vals = {
            '头像': t.get('头像', ''),
            '昵称': t.get('昵称', ''),
            '带货类目1': cat1, '带货类目2': cat2, '带货类目3': cat3,
            '评分': t.get('评分', ''),
            '粉丝数(列表)': t.get('粉丝数(列表)', ''),
            '带货销售总额': t.get('带货销售总额', ''),
            '短视频销售额': t.get('短视频销售额', ''),
            '回复率高': t.get('回复率高', ''),
            '有认证': t.get('有认证', ''),
            '可开发票': t.get('可开发票', ''),
            '达人详情链接': t.get('达人详情链接', ''),
            '总销量': t.get('总销量', ''),
            '跟买人数': t.get('跟买人数', ''),
            '回头客': t.get('回头客', ''),
            '品类占比': t.get('品类占比', ''),
            '带货销售额': t.get('带货销售额', ''),
            '客单价': t.get('客单价', ''),
            '粉丝数带货概览': t.get('粉丝数带货概览', ''),
            '直播占比': t.get('直播占比', ''),
            '粉丝性别': t.get('粉丝性别', ''),
            '粉丝年龄': t.get('粉丝年龄', ''),
            '粉丝地域': t.get('粉丝地域', ''),
            '粉丝人群类别': t.get('粉丝人群类别', ''),
            '粉丝购物偏好': t.get('粉丝购物偏好', ''),
            '购买力区间': t.get('购买力区间', ''),
            '带货渠道': t.get('带货渠道', ''),
            '直播销售额': t.get('直播销售额', ''),
            '场均成交额': t.get('场均成交额', ''),
            '场均观看人数': t.get('场均观看人数', ''),
            '总带货场次': t.get('总带货场次', ''),
            '直播明细': t.get('直播明细', ''),
            '视频销售额': t.get('视频销售额', ''),
            '条均成交额': t.get('条均成交额', ''),
            '条均点赞数': t.get('条均点赞数', ''),
            '总带货条数': t.get('总带货条数', ''),
            '短视频明细': t.get('短视频明细', ''),
            '微信号': t.get('微信号', ''),
            '手机号': t.get('手机号', ''),
            'im链接': t.get('imLink', ''),
        }
        for ci, h in enumerate(hdrs, 1):
            ws.cell(row=ri, column=ci, value=vals.get(h, '')).alignment = Alignment(
                vertical="center", wrap_text=True
            )

    for ci in range(1, len(hdrs) + 1):
        ml = max(8, len(str(hdrs[ci - 1])))
        for ri in range(2, min(len(talents) + 2, 101)):
            ml = max(ml, len(str(ws.cell(row=ri, column=ci).value or '')))
        ws.column_dimensions[get_column_letter(ci)].width = min(ml + 2, 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}{len(talents) + 1}"
    wb.save(fp)
    print(f"📄 {fp}", flush=True)
    return fp


if __name__ == '__main__':
    asyncio.run(main())
